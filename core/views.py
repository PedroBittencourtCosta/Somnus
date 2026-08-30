
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from core.Scaleprocessor import EscalaEngine, SafeParser
from core.services import calcular_e_salvar_resultados

from ethics.models import TCLE, AceiteTCLE
from .models import Questionario, Secao, Pergunta, Alternativa, RespostaQuestionario, RespostaPergunta, EscalaConfig
import json
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction

def index_view(request: HttpRequest):
    return render(request, 'home.html')

@login_required
def responder_questionario(request, pk):
    questionario = get_object_or_404(Questionario, pk=pk)
    ultimo_tcle = TCLE.objects.order_by('-versao').first()

    # --- LÓGICA TCLE: POR ATENDIMENTO (SESSÃO) ---
    tcle_aceito_na_sessao = request.session.get('tcle_aceito', False)

    if not tcle_aceito_na_sessao and ultimo_tcle:
        context = {
            'questionario': questionario,
            'exibir_tcle': True,
            'tcle': ultimo_tcle,
        }
        return render(request, 'responder_questionario.html', context)
    
    # Configuração da paginação e sessões
    secoes_list = questionario.secoes.all().order_by('ordem')
    paginator = Paginator(secoes_list, 1)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Proteção GET: impede acesso direto via URL a páginas não respondidas
    max_permitida_get = request.session.get('max_pagina_respondida', 1)
    if page_obj.number > max_permitida_get:
        return redirect(f"{request.path}?page={max_permitida_get}")

    secao_atual = page_obj.object_list[0] if page_obj.object_list else None

    if 'respostas_temp' not in request.session:
        request.session['respostas_temp'] = {}

    # Rastreia a página mais avançada já respondida nesta sessão
    if 'max_pagina_respondida' not in request.session:
        request.session['max_pagina_respondida'] = 1

    if request.method == 'POST':
        acao = request.POST.get('acao')

        # Aceite do TCLE
        if acao == 'aceitar_tcle':
            request.session['tcle_aceito'] = True
            request.session.modified = True
            return redirect(request.path)

        # Salva respostas da página atual na sessão
        if secao_atual:
            for pergunta in secao_atual.perguntas.all():
                request.session['respostas_temp'][str(pergunta.id)] = {
                    'alternativa': request.POST.get(f'pergunta_{pergunta.id}'),
                    'texto': request.POST.get(f'pergunta_{pergunta.id}_texto'),
                    'identificador': pergunta.identificador 
                }
        
        request.session.modified = True

        if acao == 'proximo' and page_obj.has_next():
            # Avançou: atualiza o máximo de página respondida
            max_atual = request.session.get('max_pagina_respondida', 1)
            if page_obj.number >= max_atual:
                request.session['max_pagina_respondida'] = page_obj.number + 1
                request.session.modified = True
            return redirect(f"{request.path}?page={page_obj.next_page_number()}")
        elif acao == 'anterior' and page_obj.has_previous():
            return redirect(f"{request.path}?page={page_obj.previous_page_number()}")
        
        elif acao and acao.startswith('ir_pagina_'):
            try:
                target_page = int(acao.replace('ir_pagina_', ''))
            except (ValueError, TypeError):
                target_page = 1
            # Clamp server-side: só permite navegar até a página máxima respondida
            max_permitida = request.session.get('max_pagina_respondida', 1)
            target_page = min(target_page, max_permitida)
            target_page = max(target_page, 1)
            return redirect(f"{request.path}?page={target_page}")
        
        elif acao == 'finalizar':
            respostas_cache = request.session.get('respostas_temp', {})
            ultimo_tcle = TCLE.objects.order_by('-versao').first()

            # 1. Extraímos o nome do paciente do cache (procurando pelo identificador técnico 'nome')
            nome_extraido = "Não informado"
            for v in respostas_cache.values():
                if v.get('identificador') == 'nome':
                    # Pega o texto preenchido na pergunta de nome
                    nome_extraido = v.get('texto') or "Não informado"
                    break

           # 2. Criamos a RespostaQuestionario (Modelo novo)
            res_quest = RespostaQuestionario.objects.create(
                pesquisadora=request.user, # Aluna logada
                questionario=questionario,
                paciente_nome=nome_extraido
            )

            # 3. NOVO: Relacionamos a resposta ao TCLE aceito nesta sessão
            if ultimo_tcle:
                AceiteTCLE.objects.create(
                    resposta_questionario=res_quest,
                    tcle=ultimo_tcle
                )
            
            # 4. Salva todas as respostas (incluindo idade, sexo, etc., que agora são perguntas)
            for p_id, valores in respostas_cache.items():
                pergunta = Pergunta.objects.get(id=p_id)
                alt = None
                if valores.get('alternativa'):
                    try:
                        alt = Alternativa.objects.get(id=valores['alternativa'])
                    except (Alternativa.DoesNotExist, ValueError):
                        alt = None
                
                RespostaPergunta.objects.create(
                    resposta_questionario=res_quest,
                    pergunta=pergunta,
                    alternativa=alt,
                    resposta_texto=valores.get('texto')
                )
            
            # 5. Calcula e persiste os resultados das escalas (cache para o dashboard)
            calcular_e_salvar_resultados(res_quest)

            # 6. Limpeza da sessão para o próximo atendimento
            del request.session['respostas_temp']
            request.session['tcle_aceito'] = False
            request.session['max_pagina_respondida'] = 1
            request.session.modified = True

            messages.success(request, f"Avaliação de {nome_extraido} concluída!")
            return redirect('home')

    context = {
        'questionario': questionario,
        'secao': secao_atual,
        'page_obj': page_obj,
        'respostas_preenchidas': request.session.get('respostas_temp', {}),
        'progresso': int((page_obj.number / paginator.num_pages) * 100),
        'max_pagina_respondida': request.session.get('max_pagina_respondida', 1),
    }
    return render(request, 'responder_questionario.html', context)

def lista_questionarios(request):
    # Exibe apenas questínarios ativos para coleta assistida.
    questionarios = Questionario.objects.filter(ativo=True).order_by('-data_criacao')
    return render(request, 'lista_questionarios.html', {
        'questionarios': questionarios,
        'respondidos': []
    })


@login_required
def gerenciar_questionarios(request):
    """Tela de gestão centralizada (pesquisadores/staff). Exibe todos, ativos e inativos."""
    questionarios = Questionario.objects.all().order_by('-data_criacao')
    return render(request, 'gerenciar_questionarios.html', {
        'questionarios': questionarios,
    })


@login_required
@require_POST
def desativar_questionario(request, pk):
    """Soft-delete: marca o questionário como inativo. Não apaga do banco."""
    questionario = get_object_or_404(Questionario, pk=pk)
    questionario.ativo = not questionario.ativo   # Toggle: ativa/desativa
    questionario.save(update_fields=['ativo'])
    return JsonResponse({
        'status': 'success',
        'ativo': questionario.ativo,
        'message': 'Questionário ativado.' if questionario.ativo else 'Questionário desativado.'
    })





@login_required
def dashboard_respostas(request):
    from collections import Counter, defaultdict
    from .models import ResultadoEscala, RespostaPergunta
    import json as _json

    # ── Filtro por questionário ──────────────────────────────────────────────
    questionario_id = request.GET.get('questionario', '')

    questionarios = Questionario.objects.filter(ativo=True).order_by('titulo')

    # Queryset base de respostas
    qs_respostas = RespostaQuestionario.objects.select_related('pesquisadora', 'questionario').all()
    if questionario_id:
        qs_respostas = qs_respostas.filter(questionario_id=questionario_id)

    total_avaliacoes = qs_respostas.count()

    # ── KPIs — calculados do cache ResultadoEscala ──────────────────────────
    def _kpi_prevalencia(escala_nome_contains, classificacoes_ruins, ids_respostas=None):
        """Retorna (n_afetados, percentual) para uma classificação de risco."""
        from django.db.models import Q
        qs = ResultadoEscala.objects.filter(escala_config__nome__icontains=escala_nome_contains)
        if ids_respostas is not None:
            qs = qs.filter(resposta_questionario_id__in=ids_respostas)
        total = qs.count()
        if not total:
            return 0, 0
            
        if isinstance(classificacoes_ruins, str):
            classificacoes_ruins = [classificacoes_ruins]
            
        query = Q()
        for c in classificacoes_ruins:
            query |= Q(classificacao__icontains=c)
            
        afetados = qs.filter(query).count()
        return afetados, round((afetados / total) * 100, 1)

    ids_filtradas = list(qs_respostas.values_list('id', flat=True)) if questionario_id else None

    n_psqi_ruim, pct_psqi_ruim       = _kpi_prevalencia('PSQI',  ['Qualidade Ruim'], ids_filtradas)
    n_ese_sed,   pct_ese_sed         = _kpi_prevalencia('ESE',   ['Sonolência Diurna Excessiva', 'Sonolência Diurna'], ids_filtradas)
    n_srq_tmc,   pct_srq_tmc         = _kpi_prevalencia('SRQ-20', ['Sofrimento mental moderado', 'Sofrimento mental grave', 'Suspeita de TMC'], ids_filtradas)
    n_k10_risco, pct_k10_risco       = _kpi_prevalencia('K10',   ['Provável transtorno'], ids_filtradas)

    # ── Dados para gráficos Chart.js ─────────────────────────────────────────
    def _distribuicao(escala_nome_contains, ids_respostas=None):
        """Retorna {classificação: contagem} para uma estratégia."""
        qs = ResultadoEscala.objects.filter(escala_config__nome__icontains=escala_nome_contains)
        if ids_respostas is not None:
            qs = qs.filter(resposta_questionario_id__in=ids_respostas)
        dist = Counter(qs.values_list('classificacao', flat=True))
        return dict(dist)

    dist_psqi  = _distribuicao('PSQI', ids_filtradas)
    dist_ese   = _distribuicao('ESE',  ids_filtradas)
    dist_k10   = _distribuicao('K10',  ids_filtradas)
    dist_srq20 = _distribuicao('SRQ-20', ids_filtradas)
    dist_audit = _distribuicao('AUDIT', ids_filtradas)

    # Dados de correlação PSQI × DASS-21 Depressão (scatter)
    scatter_psqi_dass = []
    psqi_resultados = ResultadoEscala.objects.filter(escala_config__nome__icontains='PSQI')
    if ids_filtradas is not None:
        psqi_resultados = psqi_resultados.filter(resposta_questionario_id__in=ids_filtradas)

    for r in psqi_resultados.select_related('resposta_questionario'):
        psqi_score = r.score_principal
        if psqi_score is None:
            continue
        dass_obj = ResultadoEscala.objects.filter(
            resposta_questionario=r.resposta_questionario,
            escala_config__nome__icontains='DASS-21'
        ).first()
        if dass_obj and dass_obj.resultado_json:
            depressao_score = dass_obj.resultado_json.get('dass_depressao_total')
            if depressao_score is not None:
                scatter_psqi_dass.append({
                    'x': psqi_score,
                    'y': depressao_score
                })



    context = {
        # Filtros
        'questionarios': questionarios,
        'filtro_questionario': questionario_id,

        # KPIs
        'total_avaliacoes': total_avaliacoes,
        'pct_psqi_ruim': pct_psqi_ruim,
        'n_psqi_ruim': n_psqi_ruim,
        'pct_ese_sed': pct_ese_sed,
        'n_ese_sed': n_ese_sed,
        'pct_srq_tmc': pct_srq_tmc,
        'n_srq_tmc': n_srq_tmc,
        'pct_k10_risco': pct_k10_risco,
        'n_k10_risco': n_k10_risco,

        # Dados para gráficos (serializados como JSON)
        'dist_psqi_json':  _json.dumps(dist_psqi),
        'dist_ese_json':   _json.dumps(dist_ese),
        'dist_k10_json':   _json.dumps(dist_k10),
        'dist_srq20_json': _json.dumps(dist_srq20),
        'dist_audit_json': _json.dumps(dist_audit),
        'scatter_psqi_dass_json': _json.dumps(scatter_psqi_dass),
    }
        
    return render(request, 'dashboard_respostas.html', context)

@login_required
def relatorios_medicos(request):
    from django.core.paginator import Paginator
    from .models import Questionario, RespostaQuestionario

    questionarios = Questionario.objects.filter(ativo=True).order_by('titulo')
    qs_respostas = RespostaQuestionario.objects.select_related('pesquisadora', 'questionario').all()

    questionario_id = request.GET.get('questionario', '')
    if questionario_id:
        qs_respostas = qs_respostas.filter(questionario_id=questionario_id)

    sort = request.GET.get('sort', '-data_submissao')
    valid_sorts = ['codigo_paciente', '-codigo_paciente', 'questionario__titulo', '-questionario__titulo', 'data_submissao', '-data_submissao']
    if sort in valid_sorts:
        qs_respostas = qs_respostas.order_by(sort)
    else:
        qs_respostas = qs_respostas.order_by('-data_submissao')

    # Paginação para 10 itens por página
    paginator = Paginator(qs_respostas, 10)
    page_number = request.GET.get('page')
    todas_respostas = paginator.get_page(page_number)

    context = {
        'questionarios': questionarios,
        'filtro_questionario': questionario_id,
        'todas_respostas': todas_respostas,
        'current_sort': sort,
    }
    return render(request, 'relatorios_medicos.html', context)


@login_required
@require_POST
def recalcular_escalas(request):
    """
    Força o recálculo e atualização de todos os ResultadoEscala.

    Aceita POST com campo opcional 'questionario' (ID) para limitar o
    recálculo às respostas de um questionário específico.

    Retorna JSON: { sucesso, falhas, total }
    """
    questionario_id = request.POST.get('questionario', '').strip()

    qs = RespostaQuestionario.objects.select_related('questionario').all()
    if questionario_id:
        qs = qs.filter(questionario_id=questionario_id)

    total = qs.count()
    sucesso = 0
    falhas = 0
    erros = []

    for resposta in qs.iterator():
        try:
            calcular_e_salvar_resultados(resposta)
            sucesso += 1
        except Exception as e:
            falhas += 1
            erros.append(f"[{resposta.codigo_paciente}] {e}")

    return JsonResponse({
        'sucesso': sucesso,
        'falhas': falhas,
        'total': total,
        'erros': erros,
    })


def get_answers_by_section(respostas_queryset, section_id):
    """Filtra respostas de uma seção e mapeia pela posição relativa."""
    respostas_secao = respostas_queryset.filter(
        pergunta__secao_id=section_id
    ).order_by('pergunta__ordem')
    
    return {i+1: rp.alternativa.valor for i, rp in enumerate(respostas_secao) if rp.alternativa}


@login_required
def exportar_respostas_excel(request, pk):
    # 1. BUSCA DE DADOS
    res_quest = get_object_or_404(RespostaQuestionario, pk=pk)
    pesquisadora = res_quest.pesquisadora 
    
    respostas_qs = RespostaPergunta.objects.filter(
        resposta_questionario=res_quest
    ).select_related('pergunta', 'alternativa')

    # 2. MAPEAMENTO HÍBRIDO
    answers_map = {}
    for rp in respostas_qs:
        identificador = rp.pergunta.identificador
        if identificador:
            if rp.alternativa:
                answers_map[identificador] = rp.alternativa.valor
            elif rp.resposta_texto:
                answers_map[identificador] = rp.resposta_texto

    # 3. PROCESSAMENTO DAS ESCALAS VINCULADAS
    escalas_resultados = []
    for escala in res_quest.questionario.escalas_config.all():
        resultado = EscalaEngine.processar(answers_map, escala)
        if "erro" not in resultado:
            chaves_valor = []
            chaves_status = {}
            
            for k, v in resultado.items():
                k_str = str(k)
                if k_str.endswith('_status') or k_str == 'status' or k_str.endswith('_classificacao') or k_str == 'classificacao':
                    prefix = k_str.replace('_status', '').replace('status', '').replace('_classificacao', '').replace('classificacao', '')
                    chaves_status[prefix] = v
                elif isinstance(v, (int, float, str)) and not isinstance(v, list):
                    chaves_valor.append((k_str, v))
            
            for k, v in chaves_valor:
                prefix = k.replace('_total', '').replace('_global', '').replace('_valor', '')
                
                referencia = "-"
                if prefix in chaves_status:
                    referencia = chaves_status.pop(prefix)
                elif k in chaves_status:
                    referencia = chaves_status.pop(k)
                elif "" in chaves_status:
                    referencia = chaves_status.pop("")
                    
                key_formatted = k.replace('_', ' ').title()
                escalas_resultados.append([f"{escala.nome} - {key_formatted}", v, str(referencia)])
                
            for pref, ref in chaves_status.items():
                nome_formatado = f"{escala.nome} - Status {pref.replace('_', ' ').title()}".strip()
                if nome_formatado.endswith('- Status'):
                    nome_formatado = f"{escala.nome} - Status"
                escalas_resultados.append([nome_formatado, "-", str(ref)])
    
    # IMC Fallback (caso ainda queiram sem precisar configurar)
    peso = SafeParser.to_float(answers_map.get('peso', 0))
    altura = SafeParser.to_float(answers_map.get('altura', 0))
    if altura > 0:
        imc = round(peso / (altura ** 2), 2)
        escalas_resultados.append(["IMC Calculado Automaticamente", imc, "kg/m²"])

    # 4. CRIAÇÃO DO EXCEL
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Avaliação"

    # Estilos
    azul_somnus = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
    cinza_claro = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    fonte_branca = Font(color='FFFFFF', bold=True)
    center_align = Alignment(horizontal='center', vertical='center')

    # Cabeçalho
    ws.merge_cells('A1:C1')
    ws['A1'] = "SOMNUS - RELATÓRIO TÉCNICO"
    ws['A1'].font = Font(size=14, bold=True, color='1A365D')
    ws['A1'].alignment = center_align

    ws.append(['Paciente:', res_quest.paciente_nome])
    ws.append(['Pesquisadora:', pesquisadora.username])
    ws.append(['Data:', res_quest.data_submissao.strftime('%d/%m/%Y %H:%M')])
    ws.append([])

    def adicionar_secao(titulo, dados):
        if not dados: return
        ws.append([titulo])
        ws.merge_cells(f'A{ws.max_row}:C{ws.max_row}')
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=1).fill = cinza_claro
        
        ws.append(['Escala/Métrica', 'Resultado', 'Referência'])
        for cell in ws[ws.max_row]:
            cell.fill = azul_somnus
            cell.font = fonte_branca

        for linha in dados:
            ws.append(linha)
        ws.append([])

    # Adicionando os Scores Calculados Dinamicamente
    adicionar_secao("I. RESULTADOS DAS ESCALAS", escalas_resultados)

    # II. Detalhamento
    ws.append(['IV. DETALHAMENTO DAS RESPOSTAS'])
    ws.append(['ID', 'Pergunta', 'Valor'])
    for cell in ws[ws.max_row]:
        cell.fill = azul_somnus
        cell.font = fonte_branca

    for rp in respostas_qs.order_by('pergunta__ordem'):
        valor = rp.alternativa.valor if rp.alternativa else rp.resposta_texto
        ws.append([rp.pergunta.identificador or f"ID_{rp.pergunta.id}", rp.pergunta.conteudo[:100], valor])

    # --- CORREÇÃO DO ERRO: AJUSTE DE LARGURA SEGURO ---
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i) # Pega a letra da coluna pelo índice (1=A, 2=B...)
        
        for cell in column_cells:
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    # 5. RETORNO
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Somnus_Relatorio_{res_quest.id}.xlsx"'
    wb.save(response)
    return response

@login_required
def exportar_resultados_massa_excel(request):
    questionario_id = request.GET.get('questionario')
    
    if questionario_id:
        respostas = RespostaQuestionario.objects.filter(questionario_id=questionario_id).select_related('questionario', 'pesquisadora')
    else:
        respostas = RespostaQuestionario.objects.all().select_related('questionario', 'pesquisadora')
        
    respostas = respostas.order_by('-data_submissao')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados em Massa"
    
    # 1. Cabecalhos fixos
    cabecalhos = ['ID', 'Código Paciente', 'Nome Paciente', 'Data', 'Questionário']
    
    todas_perguntas = {}
    todas_escalas = {}
    
    from django.db.models import Prefetch
    respostas = respostas.prefetch_related(
        'respostas__pergunta',
        'respostas__alternativa',
        'questionario__escalas_config',
        'resultados_escalas__escala_config'
    )
    
    dados_linhas = []
    
    for res in respostas:
        linha_base = {
            'ID': res.id,
            'Código Paciente': res.codigo_paciente,
            'Nome Paciente': res.paciente_nome,
            'Data': res.data_submissao.strftime('%d/%m/%Y %H:%M'),
            'Questionário': res.questionario.titulo_curto
        }
        
        # Respostas brutas
        for rp in res.respostas.all():
            chave = f"P: {rp.pergunta.identificador or rp.pergunta.conteudo[:30]}"
            if chave not in todas_perguntas:
                todas_perguntas[chave] = True
            
            valor = rp.alternativa.valor if rp.alternativa else rp.resposta_texto
            linha_base[chave] = valor
            
        # Resultados de escalas (usando o cache)
        for esc in res.resultados_escalas.all():
            chave_score = f"E: {esc.escala_config.nome} (Score)"
            chave_class = f"E: {esc.escala_config.nome} (Classificação)"
            
            if chave_score not in todas_escalas: todas_escalas[chave_score] = True
            if chave_class not in todas_escalas: todas_escalas[chave_class] = True
            
            linha_base[chave_score] = esc.score_principal
            linha_base[chave_class] = esc.classificacao
            
        dados_linhas.append(linha_base)
        
    cabecalhos.extend(list(todas_escalas.keys()))
    cabecalhos.extend(list(todas_perguntas.keys()))
    
    ws.append(cabecalhos)
    
    azul_somnus = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
    fonte_branca = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = azul_somnus
        cell.font = fonte_branca
        
    for linha_dict in dados_linhas:
        linha_excel = []
        for col in cabecalhos:
            linha_excel.append(linha_dict.get(col, ""))
        ws.append(linha_excel)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Somnus_Exportacao_Massa.xlsx"'
    wb.save(response)
    return response


@login_required
def editar_questionario_view(request, pk=None):
    questionario_json = "{}"
    questionario = None
    
    if pk:
        questionario = get_object_or_404(Questionario, pk=pk)
        dados = {
            "id": questionario.id,
            "titulo": questionario.titulo,
            "descricao": questionario.descricao,
            "secoes": []
        }
        for secao in questionario.secoes.all().order_by('ordem'):
            sec_data = {
                "id": secao.id,
                "titulo": secao.titulo,
                "instrucao": secao.instrucao,
                "layout": secao.layout,
                "ordem": secao.ordem,
                "perguntas": []
            }
            for perg in secao.perguntas.all().order_by('ordem'):
                perg_data = {
                    "id": perg.id,
                    "conteudo": perg.conteudo,
                    "tipo": perg.tipo,
                    "mascara": perg.mascara,
                    "config_mista": perg.config_mista,
                    "obrigatoria": perg.obrigatoria,
                    "identificador": perg.identificador,
                    "ordem": perg.ordem,
                    "depende_de_alternativa_ids": list(
                        perg.depende_de_alternativa.values_list('id', flat=True)
                    ),
                    "depende_de_texto_de_id": perg.depende_de_texto_de_id,
                    "alternativas": []
                }
                for alt in perg.alternativas.all():
                    perg_data["alternativas"].append({
                        "id": alt.id,
                        "conteudo": alt.conteudo,
                        "valor": alt.valor
                    })
                sec_data["perguntas"].append(perg_data)
            dados["secoes"].append(sec_data)
        questionario_json = json.dumps(dados)

    return render(request, 'criar_questionario.html', {
        'questionario': questionario,
        'questionario_json': questionario_json
    })

@login_required
@require_POST
def salvar_questionario_api(request):
    try:
        data = json.loads(request.body)
        questionario_id = data.get('id')
        
        with transaction.atomic():
            if questionario_id:
                quest = Questionario.objects.get(pk=questionario_id)
                quest.titulo = data.get('titulo', '')
                quest.descricao = data.get('descricao', '')
                quest.save()
            else:
                quest = Questionario.objects.create(
                    titulo=data.get('titulo', ''),
                    descricao=data.get('descricao', '')
                )
            
            secoes_ids = []
            perguntas_ids = []
            alternativas_ids = []
            
            # Mapeamentos para resolver dependências (temp_id → real_id)
            alt_id_map = {}       # "temp_xyz" ou "123" → real_id da alternativa
            perg_id_map = {}      # "temp_xyz" ou "123" → real_id da pergunta
            perguntas_com_deps = []  # [(pergunta_obj, perg_data_dict), ...]
            
            # ── PASSADA 1: Criar/Atualizar todas as seções, perguntas e alternativas ──
            for index_secao, sec_data in enumerate(data.get('secoes', [])):
                sec_id = sec_data.get('id')
                if sec_id and str(sec_id).isdigit():
                    secao = Secao.objects.get(pk=sec_id, questionario=quest)
                    secao.titulo = sec_data.get('titulo', '')
                    secao.instrucao = sec_data.get('instrucao', '')
                    secao.layout = sec_data.get('layout', 'LISTA')
                    secao.ordem = index_secao + 1
                    secao.save()
                else:
                    secao = Secao.objects.create(
                        questionario=quest,
                        titulo=sec_data.get('titulo', ''),
                        instrucao=sec_data.get('instrucao', ''),
                        layout=sec_data.get('layout', 'LISTA'),
                        ordem=index_secao + 1
                    )
                secoes_ids.append(secao.id)
                
                for index_perg, perg_data in enumerate(sec_data.get('perguntas', [])):
                    perg_id = perg_data.get('id')
                    perg_temp_id = perg_data.get('temp_id')  # ID temporário do frontend
                    
                    if perg_id and str(perg_id).isdigit():
                        perg = Pergunta.objects.get(pk=perg_id, secao__questionario=quest)
                        perg.secao = secao
                        perg.conteudo = perg_data.get('conteudo', '')
                        perg.tipo = perg_data.get('tipo', 'MC')
                        perg.mascara = perg_data.get('mascara', 'NENHUMA')
                        perg.config_mista = perg_data.get('config_mista', 'QUALQUER')
                        perg.obrigatoria = perg_data.get('obrigatoria', True)
                        perg.identificador = perg_data.get('identificador', '')
                        perg.ordem = index_perg + 1
                        perg.save()
                    else:
                        perg = Pergunta.objects.create(
                            secao=secao,
                            conteudo=perg_data.get('conteudo', ''),
                            tipo=perg_data.get('tipo', 'MC'),
                            mascara=perg_data.get('mascara', 'NENHUMA'),
                            config_mista=perg_data.get('config_mista', 'QUALQUER'),
                            obrigatoria=perg_data.get('obrigatoria', True),
                            identificador=perg_data.get('identificador', ''),
                            ordem=index_perg + 1
                        )
                    perguntas_ids.append(perg.id)
                    
                    # Registrar mapeamento de IDs de perguntas
                    if perg_id and str(perg_id).isdigit():
                        perg_id_map[str(perg_id)] = perg.id
                    if perg_temp_id:
                        perg_id_map[str(perg_temp_id)] = perg.id
                    
                    # Guardar para processar dependências na passada 2
                    perguntas_com_deps.append((perg, perg_data))
                    
                    for index_alt, alt_data in enumerate(perg_data.get('alternativas', [])):
                        alt_id = alt_data.get('id')
                        alt_temp_id = alt_data.get('temp_id')  # ID temporário do frontend
                        
                        try:
                            valor_int = int(alt_data.get('valor', 0))
                        except ValueError:
                            valor_int = 0
                            
                        if alt_id and str(alt_id).isdigit():
                            alt = Alternativa.objects.get(pk=alt_id, pergunta=perg)
                            alt.conteudo = alt_data.get('conteudo', '')
                            alt.valor = valor_int
                            alt.save()
                        else:
                            alt = Alternativa.objects.create(
                                pergunta=perg,
                                conteudo=alt_data.get('conteudo', ''),
                                valor=valor_int
                            )
                        alternativas_ids.append(alt.id)
                        
                        # Registrar mapeamento de IDs de alternativas
                        if alt_id and str(alt_id).isdigit():
                            alt_id_map[str(alt_id)] = alt.id
                        if alt_temp_id:
                            alt_id_map[str(alt_temp_id)] = alt.id
            
            # Limpar registros removidos
            Alternativa.objects.filter(pergunta__secao__questionario=quest).exclude(id__in=alternativas_ids).delete()
            Pergunta.objects.filter(secao__questionario=quest).exclude(id__in=perguntas_ids).delete()
            Secao.objects.filter(questionario=quest).exclude(id__in=secoes_ids).delete()
            
            # ── PASSADA 2: Resolver e salvar dependências ──
            for perg, perg_data in perguntas_com_deps:
                # Dependência por alternativa(s) selecionada(s)
                dep_alt_ids_raw = perg_data.get('depende_de_alternativa_ids', [])
                real_alt_ids = []
                for raw_id in dep_alt_ids_raw:
                    real_id = alt_id_map.get(str(raw_id))
                    if real_id:
                        real_alt_ids.append(real_id)
                    elif str(raw_id).isdigit():
                        # ID numérico que já existia no banco
                        real_alt_ids.append(int(raw_id))
                perg.depende_de_alternativa.set(real_alt_ids)
                
                # Dependência por texto de outra pergunta
                dep_texto_raw = perg_data.get('depende_de_texto_de_id')
                if dep_texto_raw:
                    real_perg_id = perg_id_map.get(str(dep_texto_raw))
                    if real_perg_id:
                        perg.depende_de_texto_de_id = real_perg_id
                    elif str(dep_texto_raw).isdigit():
                        perg.depende_de_texto_de_id = int(dep_texto_raw)
                    else:
                        perg.depende_de_texto_de_id = None
                else:
                    perg.depende_de_texto_de_id = None
                perg.save(update_fields=['depende_de_texto_de'])
            
        return JsonResponse({'status': 'success', 'questionario_id': quest.id})
        
    except Exception as e:
        import traceback
        return JsonResponse({'status': 'error', 'message': str(e), 'trace': traceback.format_exc()}, status=400)

@login_required
def configurar_escala_view(request, pk):
    questionario = get_object_or_404(Questionario, pk=pk)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            acao = data.get('acao')
            escala_id = data.get('escala_id')
            
            if acao == 'vincular':
                escala = get_object_or_404(EscalaConfig, id=escala_id)
                escala.questionarios.add(questionario)
                return JsonResponse({'status': 'success', 'escala_id': escala.id, 'nome': escala.nome})
            elif acao == 'desvincular':
                escala = get_object_or_404(EscalaConfig, id=escala_id)
                escala.questionarios.remove(questionario)
                return JsonResponse({'status': 'success', 'escala_id': escala.id, 'nome': escala.nome})
            elif acao == 'desativar':
                # Soft-delete: irreversível pela UI do pesquisador
                escala = get_object_or_404(EscalaConfig, id=escala_id)
                escala.ativo = False
                escala.save(update_fields=['ativo'])
                return JsonResponse({'status': 'success', 'escala_id': escala.id})
            
            nome = data.get('nome', f'Escala para {questionario.titulo}'[:100])
            
            if escala_id:
                # Editar escala existente
                config = get_object_or_404(EscalaConfig, id=escala_id)
                config.nome = nome
            else:
                # Criar nova escala
                config = EscalaConfig.objects.create(nome=nome)
                config.questionarios.add(questionario)
                
            config.strategy_class = data.get('strategy_class', 'DYNAMIC')
            if config.strategy_class == 'DYNAMIC':
                config.config_dinamica = data.get('config_dinamica', {})
            else:
                config.config_dinamica = None
                
            config.save()
            return JsonResponse({'status': 'success', 'escala_id': config.id, 'nome': config.nome})
        except Exception as e:
            import traceback
            return JsonResponse({'status': 'error', 'message': str(e), 'trace': traceback.format_exc()}, status=400)
            
    # Coleta todas as perguntas que possuem um identificador (variáveis)
    perguntas_com_id = Pergunta.objects.filter(
        secao__questionario=questionario
    ).exclude(identificador__exact='').exclude(identificador__isnull=True).values('id', 'identificador', 'conteudo')
    
    # Coleta configurações apenas das escalas ativas
    escalas_cadastradas = EscalaConfig.objects.filter(ativo=True).prefetch_related('questionarios')
    escalas_templates = []
    for e in escalas_cadastradas:
        is_linked = e.questionarios.filter(id=questionario.id).exists()
        escalas_templates.append({
            'id': e.id,
            'nome': e.nome,
            'strategy_class': e.strategy_class,
            'is_linked': is_linked,
            'config': e.config_dinamica or {}
        })
    
    context = {
        'questionario': questionario,
        'config_json': '{}',
        'perguntas_json': json.dumps(list(perguntas_com_id)),
        'escalas_templates_json': json.dumps(escalas_templates)
    }
    return render(request, 'configurar_escala.html', context)